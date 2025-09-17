import requests
import json
import os
import glob
import smtplib
import uuid
import openpyxl
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

def calculate_last_activity_timestamp(days_back=30):
    """
    Calculate timestamp for last activity filter based on days back from today.
    """
    target_date = datetime.now() - timedelta(days=days_back)
    return target_date.strftime("%Y-%m-%dT00:00:00.000Z")

def calculate_unix_timestamp_ms(days_back=30):
    """
    Calculate Unix timestamp in milliseconds for resume last updated filter.
    gte = "greater than or equal to" (profiles updated AFTER this date)
    """
    target_date = datetime.now() - timedelta(days=days_back)
    return str(int(target_date.timestamp() * 1000))

def make_indeed_request(radius, keywords, resume_updated_filter, location, language, country, offset=0, limit=50): 
    """
    Makes a POST request to Indeed's GraphQL API with hardcoded credentials and payload.
    Returns the API response.
    """
    
    url = "https://apis.indeed.com/graphql?co=DE&locale=en-DE"
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:140.0) Gecko/20100101 Firefox/140.0',
        'Accept': '*/*',
            'Accept-Language': 'en-US,en;q=0.5',
        'Accept-Encoding': 'gzip, deflate, br, zstd',
        'Referer': 'https://resumes.indeed.com/',
        'content-type': 'application/json',
        'indeed-api-key': '0f2b0de1b8ff96890172eeeba0816aaab662605e3efebbc0450745798c4b35ae',
        'indeed-ctk': '1itk4s1n0gmv5800',
        'indeed-client-sub-app': 'rezemp-discovery',
        'indeed-client-sub-app-component': './Root',
        'x-datadog-origin': 'rum',
        'x-datadog-parent-id': '2930520531491405788',  # Updated
        'x-datadog-sampling-priority': '0',
        'x-datadog-trace-id': '914046465001288166',  # Updated
        'Origin': 'https://resumes.indeed.com',
        'DNT': '1',
        'Connection': 'keep-alive',
        'Cookie': 'CTK=1itk4s1n0gmv5800; RF=wKGgxUwMHWUX5YFUwWOHJWllKOBtkHCypm6q6XELpbojtsnO9ZC_yE8zLeFvxzKc; OptanonConsent=isGpcEnabled=1^&datestamp=Mon+Jul+14+2025+09%3A04%3A20+GMT%2B0200+(Central+European+Summer+Time)^&version=202409.2.0^&browserGpcFlag=1^&isIABGlobal=false^&hosts=^&consentId=29761a2a-8bb1-4d3b-aed3-6cd19eaa08a9^&interactionCount=2^&isAnonUser=1^&landingPath=NotLandingPage^&groups=C0001%3A1%2CC0002%3A0%2CC0003%3A0%2CC0004%3A0%2CC0007%3A0^&AwaitingReconsent=false^&intType=2^&geolocation=%3B; indeed_rcc=LOCALE:CTK; cf_clearance=8Qz65PJtRDuR1lY5.nqjaBivTBgDhJpHfboNtu89KgE-1752475997-1.2.1.1-1bJdhTq5OiBglLE6uHJ0gPoLwhba7unZ43ajqkqawVhXVV1oc6w7En5btkeojOJuGw.IoAqvBIn4hKFgZ7PvT57TfDes0Pcxpi04ICYcefQ2uE_goi5JohYz.XVLo0djATpqhkvu3Ieg6oeroONYD2Ci90y.yKdRmgOg_tFLHJFuS5JD9M_jzbhsauatiEkdwy0b4b7yPXFBtT0Tredw4HmbxaCzq3b9_HMci4KL5JQ; IRF=2OjkVS608cRBJyDN80nHtHDzwzzxoxWIV0Ae_fa5Cq-OutJyB0t_LQ==; OptanonAlertBoxClosed=2025-06-13T09:59:43.706Z; CO=DE; optimizelySession=1751009392977; PPID=eyJraWQiOiI3OGFjNzFmNC04MDhjLTRjYTItOTI4NC0xYmRjNzIzMDliYzIiLCJ0eXAiOiJKV1QiLCJhbGciOiJFUzI1NiJ9.eyJzdWIiOiJhYmEyZmUwNzUyNDZlZTRhIiwibGFzdF9hdXRoX3RpbWUiOjE3NTEwMDk3NzQxOTEsImVtYWlsX3ZlcmlmaWVkIjp0cnVlLCJhdXRoIjoiZ29vZ2xlIiwiY3JlYXRlZCI6MTc0NTY3NDE4NjAwMCwiaXNzIjoiaHR0cHM6Ly9zZWN1cmUuaW5kZWVkLmNvbSIsImxhc3RfYXV0aF9sZXZlbCI6IlNUUk9ORyIsImxvZ190cyI6MTc1MTAwOTc3NDE5MSwiYXJiX2FjcHQiOmZhbHNlLCJhdWQiOiJjMWFiOGYwNGYiLCJyZW1fbWUiOnRydWUsImV4cCI6MTc1MjQ3Nzc3OCwiaWF0IjoxNzUyNDc1OTc4LCJlbWFpbCI6ImxvcmVuekBiZXlvbmRsZXZlcmFnZS5jb20ifQ.ahVONSblwrh1iUWRxbn9HEfqydzxkbOq302jK6B5PyxRZ_Dcii19hz3h0O16JRNRxyOKCzRcnH6KHhM476yd-Q; SOCK="ZaFLiaYTrg8KhpCkHQcI7ylG41k="; SHOE="6jHL9VNDfzRqUfAyKY8J-SeimanGhqPrx2q5gcRnOvOlidUf-C-7rkxJzUQNrGIlCvsaw7uTcHc5yaHRcp1anmFG4Chz9-7XwYshFq2ODQucVmOtrq9zaEANXe89h30rigDb9VfvLA928_FoTkK7egPO"; PCA=30fd7ee57925c18d; ADOC=9366375769107198; accountSwitcherPopupDismissed=true; sp=ad1ce092-29c9-4573-9959-2d84179948ef; CSRF=455404a5d36523c8dae321776edd3a82; _cfuvid=Ww3hZpbHcX47hRwgim.RoE0VLZlRjwO8cBA0ByxNuSg-1752475980326-0.0.1.1-604800000; ENC_CSRF=7yw33pQePPsFgTKDIZWhHZkRVy0bJ0jM; LOCALE=en_DE; _cfuvid=bj8Enb2uHT6qV9sDsZsKPbHWW14InxmsWR8vqvcy3Co-1752475978267-0.0.1.1-604800000; __cf_bm=9w4cJcH_AeUMVE6Pw4KltR2ax0G0v4N6i0D9sf_tDK4-1752475978-1.0.1.1-2cCyaxQo6Job2d3PLnpA_q4p9UIwSqic6qcGVCgVRn3iFQv8_HffGyXmd7PZ3k9VlPe1eUvoYfk4LXt3ugu8g06YuECA4Sd0wdNbgPE8rkQ; __cf_bm=tajrcCHwt8kDROdY0hpGJqpSuH1y5YOdJSm8Z87ZkE0-1752476660-1.0.1.1-66_Kwxt9c9V423TDc1WPIftX8qGHKB0qxOHZrfWba9TIxv1hnFsRIh50IgAc3NB.tA9cKEXpNX0YIjw5pFAkOY7Om6tQh_OLDANPxq4Iz4k; _dd_s=rum=2^&id=2f5b3be1-d20b-42e5-800b-9a8d84637b89^&created=1752476015025^&expire=1752477559533; SURF=4utXnbDO2qkDKv62rlqN5oKKTa9Ysolo',  # Updated
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-site',
        'Sec-GPC': '1',
        'Priority': 'u=4',  # Back to u=4
        'TE': 'trailers'
    }
    
    payload = {
        "operationName": "SmartSourcingResults",
        "variables": {
            "input": {
                "searchQuery": {
                    "queryString": keywords,
                    "searchLocale": {
                        "language": language,
                        "country": country
                    }
                },
                "clientSurfaceName": "sourcing-search",
                "identifiers": {
                    "jobIdentifiers": {}
                },
                "context": {
                    "overrides": {
                        "where": location,
                        "radiusInput": {
                            "distance": radius,
                            "unit": "KILOMETERS"
                        }
                    }
                },
                "filters": [
                    {
                        "key": "lastModified",
                        "afterBucket": False,
                        "gte": resume_updated_filter
                    }
                ],
                "defaultStrategyId": "RCHK8",
                "limit": limit,
                "offset": offset
            },
            "includePreferences": False,
            "includePagination": True,
            "useNewJwt": False,
            "isLoggedIn": False,
            "federateRcpRequestMetadataToSourcingProfile": True,
            "includePermissionToken": False
        },
        "extensions": {
            "remoteFragmentStats": {
                "fragmentCount": 8,
                "elapsedTime": 803
            }
        },
        "query": "query SmartSourcingResults($input: OrchestrationMatchesInput!, $includePreferences: Boolean!, $includePagination: Boolean!, $useNewJwt: Boolean!, $isLoggedIn: Boolean!, $federateRcpRequestMetadataToSourcingProfile: Boolean!, $includePermissionToken: Boolean!) {\n  findRCPMatches(input: $input) {\n    rcpRequestId\n    overallMatchCountWithoutLimit\n    searchSessionId @include(if: $includePagination)\n    overallMatchCount\n    matchConnection {\n      pageInfo @include(if: $includePagination) {\n        hasNextPage\n        hasPreviousPage\n        __typename\n      }\n      matches {\n        matchId {\n          id\n          __typename\n        }\n        highlights {\n          text\n          __typename\n        }\n        provider\n        sourcingProfile {\n          accountKey\n          permissionToken @include(if: $includePermissionToken)\n          profileCard {\n            ...ProfileCard\n            __typename\n          }\n          engagementSummary @include(if: $isLoggedIn) {\n            ...EngagementSummary\n            __typename\n          }\n          requestMetadata @include(if: $federateRcpRequestMetadataToSourcingProfile) {\n            isSegmentDecorated\n            rcpRequestId\n            __typename\n          }\n          ...TDRM_CandidateName_SourcingProfile\n          ...TDRM_CandidateLocation_SourcingProfile\n          ...TDRM_CandidateWorkExperience_SourcingProfile\n          ...TDRM_CandidateEducation_SourcingProfile\n          ...TDRM_CandidateLicensesCertifications_SourcingProfile\n          ...TDRM_CandidateSkills_SourcingProfile\n          ...TDRM_CandidateTags_SourcingProfile @include(if: $isLoggedIn)\n          ...TDRM_ProviderTags_SourcingProfile @include(if: $isLoggedIn)\n          __typename\n        }\n        __typename\n      }\n      __typename\n    }\n    refinementBuckets {\n      key\n      label\n      description\n      items {\n        count\n        label\n        value\n        __typename\n      }\n      __typename\n    }\n    searchQueryMetadata {\n      stringCorrections {\n        correction\n        original\n        __typename\n      }\n      __typename\n    }\n    __typename\n  }\n}\n\nfragment ProfileCard on SourcingProfileCard {\n  accountKey\n  activityData {\n    lastActivityTimestamp\n    __typename\n  }\n  canReceiveSMS\n  dateCreated\n  dateModified\n  educations {\n    school\n    degree\n    fromDate\n    toDate\n    __typename\n  }\n  experiences {\n    title\n    company\n    fromDate\n    toDate\n    __typename\n  }\n  firstName\n  lastClickedDate\n  lastName\n  locale\n  location {\n    localizedValue\n    __typename\n  }\n  phoneNumber\n  skills {\n    text\n    __typename\n  }\n  credentials {\n    title\n    __typename\n  }\n  preferences @include(if: $includePreferences) {\n    classLabel\n    label\n    __typename\n  }\n  resumeType\n  isFreeToContact\n  jwtTokenForResumeView @include(if: $useNewJwt)\n  __typename\n}\n\nfragment EngagementSummary on TalentEngagementSummary {\n  permissions {\n    action\n    status\n    __typename\n  }\n  __typename\n}\n\nfragment TDRM_CandidateName_SourcingProfile on SourcingProfile {\n  accountKey\n  profileCard {\n    accountKey\n    firstName\n    lastName\n    __typename\n  }\n  __typename\n}\n\nfragment TDRM_CandidateLocation_SourcingProfile on SourcingProfile {\n  accountKey\n  profileCard {\n    location {\n      localizedValue\n      __typename\n    }\n    __typename\n  }\n  __typename\n}\n\nfragment TDRM_CandidateWorkExperience_SourcingProfile on SourcingProfile {\n  accountKey\n  profileCard {\n    experiences {\n      title\n      company\n      fromDate\n      toDate\n      __typename\n    }\n    __typename\n  }\n  __typename\n}\n\nfragment TDRM_CandidateEducation_SourcingProfile on SourcingProfile {\n  accountKey\n  profileCard {\n    accountKey\n    educations {\n      school\n      degree\n      __typename\n    }\n    __typename\n  }\n  __typename\n}\n\nfragment TDRM_CandidateLicensesCertifications_SourcingProfile on SourcingProfile {\n  accountKey\n  permissionToken\n  profileCard {\n    accountKey\n    credentials {\n      title\n      __typename\n    }\n    __typename\n  }\n  __typename\n}\n\nfragment TDRM_CandidateSkills_SourcingProfile on SourcingProfile {\n  accountKey\n  permissionToken\n  profileCard {\n    accountKey\n    skills {\n      text\n      __typename\n    }\n    __typename\n  }\n  __typename\n}\n\nfragment TDRM_CandidateTags_SourcingProfile on SourcingProfile {\n  accountKey\n  profileCard {\n    dateModified\n    lastClickedDate\n    activityData {\n      lastActivityTimestamp\n      __typename\n    }\n    isFreeToContact\n    canReceiveSMS\n    __typename\n  }\n  projectBriefsConnection {\n    projectBriefs {\n      id\n      key\n      isDefault\n      name\n      __typename\n    }\n    totalCount\n    __typename\n  }\n  __typename\n}\n\nfragment TDRM_ProviderTags_SourcingProfile on SourcingProfile {\n  accountKey\n  profileCard {\n    accountKey\n    resumeType\n    __typename\n  }\n  __typename\n}\n"
    }
    
    try:
        response = requests.post(url, headers=headers, json=payload)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error making request: {e}")
        return None

def extract_candidate_data(response):
    """
    Extract specific fields from the API response.
    """
    candidates = []
    
    try:
        matches = response['data']['findRCPMatches']['matchConnection']['matches']
        
        for match in matches:
            candidate = {}
            
            # Extract match ID
            candidate['matchId'] = match['matchId']['id'] if match.get('matchId') and match['matchId'].get('id') else 'not found'
            
            # Extract profile card data
            profile_card = match['sourcingProfile']['profileCard']
            candidate['firstName'] = profile_card.get('firstName', 'not found')
            candidate['lastName'] = profile_card.get('lastName', 'not found')
            
            # Extract location
            location = profile_card.get('location', {})
            candidate['location'] = location.get('localizedValue', 'not found') if location else 'not found'
            
            # Extract experiences
            experiences = profile_card.get('experiences', [])
            candidate['experiences'] = experiences if experiences else 'not found'
            
            # Extract skills
            skills = profile_card.get('skills', [])
            candidate['skills'] = skills if skills else 'not found'
            
            # Extract educations
            educations = profile_card.get('educations', [])
            candidate['educations'] = educations if educations else 'not found'
            
            # Extract credentials
            credentials = profile_card.get('credentials', [])
            candidate['credentials'] = credentials if credentials else 'not found'
            
            # Extract permanent profile link using accountKey only
            account_key = profile_card.get('accountKey', '')
            
            if account_key:
                # Use permanent link format with accountKey
                candidate['profileLink'] = f"https://resumes.indeed.com/resume/{account_key}"
            else:
                candidate['profileLink'] = 'not found'
            
            candidates.append(candidate)
            
    except (KeyError, TypeError) as e:
        print(f"Error extracting data: {e}")
        return []
    
    return candidates

def fetch_all_candidates_with_pagination(radius, keywords, resume_updated_filter, location, language, country, max_candidates=500, params=None):
    """
    Fetch ALL available candidates using pagination until pages are exhausted.
    Prioritizes exhausting pagination before radius expansion.
    """
    all_candidates = []
    offset = 0
    page = 1
    
    # Hardcoded pagination parameters
    limit_per_page = 50
    max_pages_per_radius = 20
    
    print(f"📄 Paginating through all candidates at {radius}km radius...")
    
    while True:  # Continue until no more pages
        print(f"  → Page {page} (offset: {offset})")
        
        result = make_indeed_request(radius, keywords, resume_updated_filter, location, language, country, offset, limit_per_page)
        
        if not result:
            print("  ❌ Failed to get response from Indeed API")
            break
            
        # Extract candidates from this page
        page_candidates = extract_candidate_data(result)
        
        if not page_candidates:
            print("  ⚠️ No candidates found on this page")
            break
            
        all_candidates.extend(page_candidates)
        print(f"  ✅ Page {page}: +{len(page_candidates)} candidates (Total: {len(all_candidates)})")
        
        # Check if there are more pages - THIS IS THE KEY CHECK
        try:
            has_next_page = result['data']['findRCPMatches']['matchConnection']['pageInfo']['hasNextPage']
            if not has_next_page:
                print(f"  🏁 No more pages available at {radius}km radius")
                break
        except (KeyError, TypeError):
            print("  ⚠️ Could not determine if more pages exist")
            break
            
        # Safety check: if we got fewer than limit_per_page candidates, probably last page
        if len(page_candidates) < limit_per_page:
            print(f"  🏁 Received {len(page_candidates)} candidates (less than {limit_per_page}), likely last page")
            break
            
        # Move to next page
        offset += limit_per_page
        page += 1
        
        # Safety limit to prevent infinite loops
        if page > max_pages_per_radius:
            print(f"  ⚠️ Reached maximum page limit ({max_pages_per_radius} pages) for {radius}km radius")
            break
            
        # Stop if we have enough candidates (but let current radius finish)
        if len(all_candidates) >= max_candidates:
            print(f"  ℹ️ Reached {max_candidates} candidates, finishing this radius...")
            break
    
    print(f"📊 Pagination complete for {radius}km: {len(all_candidates)} candidates found")
    return all_candidates

def anonymize_names(candidates):
    """
    Anonymize candidate names by using matchId as identifier.
    """
    anonymized = []
    
    for candidate in candidates:
        anonymized_candidate = candidate.copy()
        
        # Use matchId as the identifier instead of real names
        match_id = candidate.get('matchId', 'Unknown')
        anonymized_candidate['firstName'] = f"Candidate_{match_id}"
        anonymized_candidate['lastName'] = ""
        
        anonymized.append(anonymized_candidate)
    
    return anonymized

def get_latest_json_file():
    """
    Find the latest response_*.json file in the current directory.
    """
    json_files = glob.glob("response_*.json")
    if not json_files:
        return None
    
    # Sort by modification time and return the newest
    latest_file = max(json_files, key=os.path.getmtime)
    return latest_file

def query_mistral_ai_batch(candidates_batch, api_key, system_prompt, user_prompt):
    """
    Query MistralAI API to filter a batch of candidates.
    """
    url = "https://api.mistral.ai/v1/chat/completions"
    
    # Build the complete prompt with candidate data
    complete_prompt = f"""CANDIDATE DATA: {json.dumps(candidates_batch, indent=2)}

{user_prompt}"""

    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }
    
    payload = {
        "model": "mistral-small-latest",
        "messages": [
            {
                "role": "system",
                "content": system_prompt
            },
            {
                "role": "user",
                "content": complete_prompt
            }
        ],
        "temperature": 0,
        "max_tokens": 1000
    }
    
    try:
        response = requests.post(url, headers=headers, json=payload)
        response.raise_for_status()
        result = response.json()
        return result['choices'][0]['message']['content']
    except requests.exceptions.RequestException as e:
        print(f"Error querying MistralAI: {e}")
        return None

def query_mistral_ai(candidates_data, output_filename, api_key, system_prompt, user_prompt):
    """
    Query MistralAI API to filter candidates in batches.
    Save results immediately to prevent data loss.
    """
    batch_size = 10  # Process 10 candidates at a time
    all_filtered_urls = []
    
    print(f"Processing {len(candidates_data)} candidates in batches of {batch_size}")
    
    # Create the output file
    with open(output_filename, 'w', encoding='utf-8') as f:
        f.write("# Qualified Candidates - Updated in Real-time\n")
        f.write("# Generated by MistralAI\n\n")
    
    for i in range(0, len(candidates_data), batch_size):
        batch = candidates_data[i:i + batch_size]
        batch_num = (i // batch_size) + 1
        total_batches = (len(candidates_data) + batch_size - 1) // batch_size
        
        print(f"Processing batch {batch_num}/{total_batches} ({len(batch)} candidates)")
        
        batch_response = query_mistral_ai_batch(batch, api_key, system_prompt, user_prompt)
        
        if batch_response:
            # Process batch response
            batch_urls = process_mistral_response(batch_response)
            if batch_urls:
                all_filtered_urls.extend(batch_urls)
                print(f"Batch {batch_num}: Found {len(batch_urls)} qualifying candidates")
                
                # Save results immediately
                with open(output_filename, 'a', encoding='utf-8') as f:
                    f.write(f"# Batch {batch_num} results:\n")
                    for url in batch_urls:
                        f.write(f"{url}\n")
                    f.write("\n")
                
                print(f"Batch {batch_num}: Results saved to {output_filename}")
            else:
                print(f"Batch {batch_num}: No qualifying candidates found")
        else:
            print(f"Batch {batch_num}: Failed to get response")
    
    return all_filtered_urls

def process_mistral_response(mistral_response):
    """
    Process MistralAI response and extract profile links.
    """
    try:
        # Clean the response
        response_clean = mistral_response.strip()
        
        # Try to parse JSON response directly
        if response_clean.startswith('[') and response_clean.endswith(']'):
            return json.loads(response_clean)
        
        # Try to extract JSON array from the response
        import re
        json_patterns = [
            r'\[.*?\]',  # Match any array
            r'\[".*?"\]',  # Match string array
            r'\[.*?https?://.*?\]',  # Match array with URLs
        ]
        
        for pattern in json_patterns:
            json_match = re.search(pattern, response_clean, re.DOTALL)
            if json_match:
                try:
                    return json.loads(json_match.group(0))
                except:
                    continue
        
        # If no JSON found, look for URLs directly
        url_pattern = r'https://resumes\.indeed\.com/resume/[^\s"\]]*'
        urls = re.findall(url_pattern, response_clean)
        if urls:
            print(f"Found {len(urls)} URLs directly from response")
            return urls
        
        print("Could not extract JSON or URLs from MistralAI response")
        print(f"Response was: {response_clean}")
        return []
        
    except json.JSONDecodeError as e:
        print(f"Error parsing MistralAI response: {e}")
        return []

def filter_candidates_with_mistral(search_keywords, location, radius, params=None):
    """
    Main function to filter candidates using MistralAI.
    """
    # Load MistralAI API key from environment variable
    api_key = os.getenv("MISTRAL_API_KEY", "83pVv0mVbOBUwSRmoPBaWg6UUkNZunTP")
    
    if not api_key:
        api_key = input("Please enter your MistralAI API key: ").strip()
        if not api_key:
            print("Error: MistralAI API key is required")
            return
    
    # Find latest JSON file
    latest_file = get_latest_json_file()
    if not latest_file:
        print("No JSON files found to process")
        return
    
    print(f"Processing file: {latest_file}")
    
    # Load candidates data
    try:
        with open(latest_file, 'r', encoding='utf-8') as f:
            candidates = json.load(f)
    except Exception as e:
        print(f"Error loading JSON file: {e}")
        return
    
    # Anonymize names using matchId
    anonymized_candidates = anonymize_names(candidates)
    
    # Create timestamp for filename
    timestamp = datetime.now().strftime("%d_%m_%Y_%H_%M_%S")
    filename = f"filtered_contacts_mistral_ai_{timestamp}.txt"
    
    # Get prompts from parameters (configurable)
    system_prompt = params.get('system_prompt', 'You are an experienced recruiter specializing in fashion retail sales positions. Your task is to evaluate candidate CVs against specific criteria and return only the profile URLs of candidates who meet ALL requirements.')
    user_prompt = params.get('user_prompt', 'EVALUATE each candidate against these EXACT criteria:...')
    
    # Query MistralAI with real-time saving
    print("Querying MistralAI for candidate filtering...")
    filtered_links = query_mistral_ai(anonymized_candidates, filename, api_key, system_prompt, user_prompt)
    
    print(f"\nFiltering complete!")
    print(f"Results saved to: {filename}")
    print(f"Total qualified candidates found: {len(filtered_links)}")
    
    if len(filtered_links) == 0:
        print("No candidates met the filtering criteria.")
    else:
        # Send email with the filtered results
        send_email_with_attachment(filename, len(filtered_links), search_keywords, location, radius, params)
        
        # Create feedback tracking file
        create_feedback_log(filename, filtered_links)

def create_feedback_log(filename, candidate_urls):
    """
    Create a feedback log file to track unsuitable profiles for prompt improvement.
    """
    timestamp = datetime.now().strftime("%d_%m_%Y_%H_%M_%S")
    feedback_filename = f"feedback_log_{timestamp}.json"
    
    feedback_data = {
        "search_timestamp": timestamp,
        "original_file": filename,
        "total_candidates": len(candidate_urls),
        "candidates": []
    }
    
    for i, url in enumerate(candidate_urls, 1):
        account_key = url.split('/')[-1] if '/' in url else f"candidate_{i}"
        feedback_data["candidates"].append({
            "candidate_id": account_key,
            "profile_url": url,
            "feedback_status": "pending",
            "feedback_timestamp": None,
            "feedback_reason": None
        })
    
    # Save feedback log
    with open(feedback_filename, 'w', encoding='utf-8') as f:
        json.dump(feedback_data, f, indent=2, ensure_ascii=False)
    
    print(f"Feedback tracking file created: {feedback_filename}")

def read_search_parameters_from_excel():
    """
    Read search parameters from Excel 'Search Parameters' sheet.
    Returns a dictionary with all search parameters.
    """
    excel_file = "SearchRecords.xlsx"
    
    try:
        if not os.path.exists(excel_file):
            print("SearchRecords.xlsx not found. Creating with default parameters...")
            create_default_search_parameters_sheet()
        
        workbook = openpyxl.load_workbook(excel_file)
        
        # Check if 'Search Parameters' sheet exists
        if 'Search Parameters' not in workbook.sheetnames:
            print("'Search Parameters' sheet not found. Creating with default parameters...")
            create_default_search_parameters_sheet()
            workbook = openpyxl.load_workbook(excel_file)
        
        worksheet = workbook['Search Parameters']
        
        # Read parameters from the sheet (assuming format: Parameter Name | Value)
        parameters = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            if row[0] and row[1]:  # Both parameter name and value exist
                param_name = str(row[0]).strip()
                param_value = row[1]
                parameters[param_name] = param_value
        
        workbook.close()
        return parameters
        
    except Exception as e:
        print(f"Error reading search parameters from Excel: {e}")
        print("Using default hardcoded parameters...")
        return get_default_parameters()

def create_default_search_parameters_sheet():
    """
    Create a default 'Search Parameters' sheet with all configurable parameters.
    """
    excel_file = "SearchRecords.xlsx"
    
    try:
        # Load existing workbook or create new one
        if os.path.exists(excel_file):
            workbook = openpyxl.load_workbook(excel_file)
        else:
            workbook = openpyxl.Workbook()
            # Remove default sheet if it exists
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
        
        # Create or get 'Search Parameters' sheet
        if 'Search Parameters' in workbook.sheetnames:
            worksheet = workbook['Search Parameters']
        else:
            worksheet = workbook.create_sheet('Search Parameters')
        
        # Clear existing content
        worksheet.delete_rows(1, worksheet.max_row)
        
        # Add headers
        worksheet.cell(row=1, column=1, value="Parameter Name")
        worksheet.cell(row=1, column=2, value="Value")
        worksheet.cell(row=1, column=3, value="Description")
        
        # Add default parameters - match actual Excel sheet fields
        default_params = [
            ("search_keywords", "Verkäufer or Verkäuferin or salesperson or Einzelhandel or Verkauf or kaufmann or kauffrau or verkaufsberater or kundenberater or verkaufsberaterin or kundenberaterin", "Keywords to search for candidates"),
            ("location", "Buxtehude", "Search location"),
            ("resume_last_updated_days", 30, "Filter profiles updated within X days"),
            ("target_candidates", 100, "Minimum candidates to find before stopping"),
            ("max_radius", 50, "Maximum search radius in km"),
            ("recipient_email", "parth@beyondleverage.com", "Email recipient for results"),
            ("user_prompt", "EVALUATE each candidate against these EXACT criteria:\n\nCRITERION 1 - Experience Duration:\nPASS: Based on the experience provided, the candidate should have atleast 4+ years total professional experience (This is a non-negotiable requirement)\nFAIL: Less than 4 years total experience\n\nCRITERION 2 - Job Stability:\nPASS: 0-1 positions shorter than 6 months in last 2 years (2023-2025)\nFAIL: 2+ positions shorter than 6 months in last 2 years\n\nCRITERION 3 - Consultative Sales Experience:\nPASS: 1+ year in advisory/consultative sales roles:\n  - Fashion retail with customer styling/advice\n  - Furniture sales with design consultation\n  - Pet supplies with animal care advice\n  - Optics/eyewear with vision consultation\n  - Electronics with technical consultation\n  - Similar customer advisory positions\n\nFAIL: Only cashier, warehouse, or basic sales without consultation\n\nREQUIRED OUTPUT FORMAT:\n[\"https://resumes.indeed.com/resume/abc123\", \"https://resumes.indeed.com/resume/def456\"]\n\nIF NO CANDIDATES QUALIFY:\n[]\n\nNo SUMMARY. NO EXPLANATION. ANALYZE AND RETURN ONLY JSON ARRAY NOW:", "User prompt for MistralAI candidate evaluation"),
            ("system_prompt", "You are an experienced recruiter specializing in fashion retail sales positions. Your task is to evaluate candidate CVs against specific criteria and return only the profile URLs of candidates who meet ALL requirements.", "System prompt for MistralAI")
        ]
        
        for i, (param_name, param_value, description) in enumerate(default_params, 2):
            worksheet.cell(row=i, column=1, value=param_name)
            worksheet.cell(row=i, column=2, value=param_value)
            worksheet.cell(row=i, column=3, value=description)
        
        # Save the workbook
        workbook.save(excel_file)
        print(f"Default search parameters created in Excel: {excel_file}")
        
    except Exception as e:
        print(f"Error creating default search parameters sheet: {e}")

def get_default_parameters():
    """
    Return default parameters as fallback - match actual Excel sheet fields.
    """
    return {
        "search_keywords": "Verkäufer or Verkäuferin or salesperson or Einzelhandel or Verkauf or kaufmann or kauffrau or verkaufsberater or kundenberater or verkaufsberaterin or kundenberaterin",
        "location": "Buxtehude",
        "resume_last_updated_days": 30,
        "target_candidates": 100,
        "max_radius": 50,
        "recipient_email": "parth@beyondleverage.com",
        "user_prompt": "EVALUATE each candidate against these EXACT criteria:\n\nCRITERION 1 - Experience Duration:\nPASS: Based on the experience provided, the candidate should have atleast 4+ years total professional experience (This is a non-negotiable requirement)\nFAIL: Less than 4 years total experience\n\nCRITERION 2 - Job Stability:\nPASS: 0-1 positions shorter than 6 months in last 2 years (2023-2025)\nFAIL: 2+ positions shorter than 6 months in last 2 years\n\nCRITERION 3 - Consultative Sales Experience:\nPASS: 1+ year in advisory/consultative sales roles:\n  - Fashion retail with customer styling/advice\n  - Furniture sales with design consultation\n  - Pet supplies with animal care advice\n  - Optics/eyewear with vision consultation\n  - Electronics with technical consultation\n  - Similar customer advisory positions\n\nFAIL: Only cashier, warehouse, or basic sales without consultation\n\nREQUIRED OUTPUT FORMAT:\n[\"https://resumes.indeed.com/resume/abc123\", \"https://resumes.indeed.com/resume/def456\"]\n\nIF NO CANDIDATES QUALIFY:\n[]\n\nNo SUMMARY. NO EXPLANATION. ANALYZE AND RETURN ONLY JSON ARRAY NOW:",
        "system_prompt": "You are an experienced recruiter specializing in fashion retail sales positions. Your task is to evaluate candidate CVs against specific criteria and return only the profile URLs of candidates who meet ALL requirements."
    }

def log_search_to_excel(search_id, keywords, location, radius, user_prompt="", system_prompt=""):
    """
    Log search details to Excel file with unique search ID.
    Uses the same file as Search Parameters but different sheet.
    """
    excel_file = "SearchRecords.xlsx"
    
    try:
        # Load existing workbook or create new one
        if os.path.exists(excel_file):
            workbook = openpyxl.load_workbook(excel_file)
        else:
            workbook = openpyxl.Workbook()
            # Remove default sheet if it exists
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
        
        # Check if Search History sheet exists
        if 'Search History' not in workbook.sheetnames:
            worksheet = workbook.create_sheet('Search History')
            # Add headers for new sheet
            headers = ["SearchID", "Keywords", "Location", "Radius", "User Prompt", "System Prompt"]
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)
        else:
            worksheet = workbook['Search History']
        
        # Find the next empty row
        next_row = worksheet.max_row + 1
        
        # Add search data
        worksheet.cell(row=next_row, column=1, value=search_id)
        worksheet.cell(row=next_row, column=2, value=keywords)
        worksheet.cell(row=next_row, column=3, value=location)
        worksheet.cell(row=next_row, column=4, value=radius)
        worksheet.cell(row=next_row, column=5, value=user_prompt)
        worksheet.cell(row=next_row, column=6, value=system_prompt)
        
        # Save the workbook
        workbook.save(excel_file)
        print(f"Search logged to Excel: {search_id}")
        
    except Exception as e:
        print(f"Error logging to Excel: {e}")

def send_email_with_attachment(filename, candidate_count, search_keywords, location, radius, params=None):
    """
    Send email with the filtered candidates file as attachment using HTML template.
    """
    # Email configuration - load from environment variables
    sender_email = os.getenv("EMAIL_USER", "aauxilliary4@gmail.com")
    sender_password = os.getenv("EMAIL_PASS", "kxoc ajnf pked zhwp")
    recipient_email = params.get('recipient_email', 'parth@beyondleverage.com') if params else 'parth@beyondleverage.com'
    
    # Read the filtered candidates URLs from the file
    candidate_urls = []
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            for line in lines:
                line = line.strip()
                if line.startswith('https://resumes.indeed.com/resume/'):
                    candidate_urls.append(line)
    except Exception as e:
        print(f"Error reading candidate URLs: {e}")
        candidate_urls = []
    
    # Generate candidate table rows with feedback buttons
    candidate_rows = ""
    for i, url in enumerate(candidate_urls, 1):
        # Extract account key from URL for tracking
        account_key = url.split('/')[-1] if '/' in url else f"candidate_{i}"
        feedback_url = f"https://matchtrex-feedback.example.com/feedback?profile={account_key}&status=unsuitable&job_id=JOB_001"
        
        candidate_rows += f"""
        <tr>
            <td style="padding: 10px; border: 1px solid #ddd; text-align: center;">{i}</td>
            <td style="padding: 10px; border: 1px solid #ddd;">
                <div style="display: flex; justify-content: space-between; align-items: center;">
                    <a href="{url}" style="color: #0066cc; text-decoration: none; font-weight: 500;">
                        👤 Kandidat {i} - Verkäufer/Einzelhandel
                    </a>
                    <a href="{feedback_url}" style="background: #dc3545; color: white; padding: 5px 10px; border-radius: 3px; text-decoration: none; font-size: 12px; margin-left: 10px;">
                        ❌ Nicht passend
                    </a>
                </div>
            </td>
        </tr>"""
    
    # Generate current timestamp
    current_time = datetime.now().strftime("%d. %B %Y, %H:%M Uhr")
    
    # Create HTML email body
    html_body = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
</head>
<body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333; max-width: 600px; margin: 0 auto; padding: 20px;">
    
    <h2 style="color: #0066cc; border-bottom: 2px solid #0066cc; padding-bottom: 10px;">
        🎯 Neue Kandidaten-Shortlist
    </h2>
    
    <p>Hallo MatchTrex Team,</p>
    
    <p>der Sourcingbot hat eine neue Kandidaten-Shortlist für euch erstellt:</p>
    
    <table style="width: 100%; border-collapse: collapse; margin: 20px 0; background: #f9f9f9; border: 1px solid #ddd;">
        <tr style="background: #0066cc; color: white;">
            <td style="padding: 12px; font-weight: bold;">Suchauftrag</td>
            <td style="padding: 12px; font-weight: bold;">Verkäufer {location}</td>
        </tr>
        <tr>
            <td style="padding: 10px; border-bottom: 1px solid #ddd;">Suchbereich</td>
            <td style="padding: 10px; border-bottom: 1px solid #ddd;">{location} + {radius}km Radius</td>
        </tr>
        <tr>
            <td style="padding: 10px; border-bottom: 1px solid #ddd;">Keywords</td>
            <td style="padding: 10px; border-bottom: 1px solid #ddd;">Verkäufer, Verkäuferin, Einzelhandel, Kundenberatung</td>
        </tr>
        <tr>
            <td style="padding: 10px; border-bottom: 1px solid #ddd;">Letzte Aktivität</td>
            <td style="padding: 10px; border-bottom: 1px solid #ddd;">Letzten 30 Tage</td>
        </tr>
        <tr>
            <td style="padding: 10px; border-bottom: 1px solid #ddd;">Gefundene Kandidaten</td>
            <td style="padding: 10px; border-bottom: 1px solid #ddd;"><strong>{candidate_count} relevante Profile</strong></td>
        </tr>
    </table>
    
    <h3 style="color: #0066cc; margin-top: 30px;">📋 Shortlist der Kandidaten:</h3>
    
    <div style="background: #f0f7ff; padding: 15px; border-radius: 5px; margin-bottom: 20px;">
        <p style="margin: 0; font-size: 14px; color: #666;">
            <strong>Hinweis:</strong> Die folgenden Kandidaten haben die KI-Bewertung erfolgreich bestanden und erfüllen die definierten Kriterien.
        </p>
    </div>
    
    <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
        <tr style="background: #f5f5f5;">
            <td style="padding: 10px; border: 1px solid #ddd; font-weight: bold; width: 50px;">#</td>
            <td style="padding: 10px; border: 1px solid #ddd; font-weight: bold;">Kandidat Profil & Feedback</td>
        </tr>
        {candidate_rows}
    </table>
    
    <div style="background: #f8f9fa; padding: 15px; border-radius: 5px; margin: 20px 0;">
        <h4 style="color: #495057; margin-top: 0;">📊 Feedback-System:</h4>
        <p style="margin: 5px 0; font-size: 13px; color: #6c757d;">
            • Klickt auf "❌ Nicht passend" wenn ein Kandidat die Anforderungen nicht erfüllt<br>
            • Das Feedback wird automatisch gespeichert und hilft bei der Verbesserung der KI-Bewertung<br>
            • Admin erhält eine Zusammenfassung aller Feedbacks zur Prompt-Optimierung
        </p>
    </div>
    
    <div style="background: #e8f5e8; padding: 15px; border-radius: 5px; margin: 20px 0;">
        <p style="margin: 0; font-size: 14px; color: #155724;">
            <strong>✅ Zusammenfassung:</strong> {candidate_count} qualifizierte Kandidaten durch KI-Bewertung identifiziert
        </p>
    </div>
    
    <h3 style="color: #0066cc;">📝 Nächste Schritte:</h3>
    <ul style="color: #666; font-size: 14px;">
        <li>Klickt auf die Links, um die vollständigen Profile zu öffnen</li>
        <li>Prüft die Kandidaten auf Passgenauigkeit</li>
        <li>Kontaktiert interessante Kandidaten direkt über Indeed</li>
        <li>Bei Fragen zum Sourcingbot wendet euch an das Tech-Team</li>
    </ul>
    
    <div style="background: #fff3cd; padding: 15px; border-radius: 5px; margin: 20px 0;">
        <p style="margin: 0; font-size: 12px; color: #856404;">
            <strong>⚠️ Wichtiger Hinweis:</strong> Diese Shortlist wird automatisch nach 6 Monaten gelöscht (DSGVO-Konformität).
        </p>
    </div>
    
    <hr style="border: none; border-top: 1px solid #ddd; margin: 30px 0;">
    
    <p style="font-size: 12px; color: #666; text-align: center;">
        Automatisch generiert durch MatchTrex Sourcingbot<br>
        Erstellt am: <strong>{current_time}</strong><br>
        Suchauftrag ID: <strong>JOB_001</strong>
    </p>
    
    <p style="font-size: 11px; color: #999; text-align: center; margin-top: 20px;">
        Diese E-Mail enthält nur Links zu öffentlich verfügbaren Profilen. Keine persönlichen Daten werden gespeichert.
    </p>
    
</body>
</html>"""
    
    # Create message
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = recipient_email
    msg['Subject'] = f"🎯 Neue Kandidaten-Shortlist - {candidate_count} qualifizierte Profile ({location})"
    
    msg.attach(MIMEText(html_body, 'html'))
    
    # Attach the file
    try:
        with open(filename, "rb") as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename= {os.path.basename(filename)}'
            )
            msg.attach(part)
    except Exception as e:
        print(f"Error attaching file: {e}")
        return False
    
    # Send email
    try:
        smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
        smtp_port = int(os.getenv("SMTP_PORT", "587"))
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, sender_password)
        text = msg.as_string()
        server.sendmail(sender_email, recipient_email, text)
        server.quit()
        print(f"Email sent successfully to {recipient_email}")
        return True
    except Exception as e:
        print(f"Error sending email: {e}")
        return False

if __name__ == "__main__":
    # Generate unique search ID
    search_id = str(uuid.uuid4())[:8]  # Short UUID for readability
    
    # Read search parameters from Excel
    print("📊 Reading search parameters from Excel...")
    params = read_search_parameters_from_excel()
    
    # Extract search parameters from Excel with proper type conversion
    try:
        search_keywords = str(params.get('search_keywords', 'Verkäufer or Verkäuferin or salesperson or Einzelhandel or Verkauf or kaufmann or kauffrau or verkaufsberater or kundenberater or verkaufsberaterin or kundenberaterin'))
        location = str(params.get('location', 'Buxtehude'))
        resume_last_updated_days = int(params.get('resume_last_updated_days', 30))
        target_candidates = int(params.get('target_candidates', 100))
        max_radius = int(params.get('max_radius', 50))
    except (ValueError, TypeError) as e:
        print(f"❌ Error converting Excel parameters: {e}")
        print("Using default parameters instead...")
        search_keywords = 'Verkäufer or Verkäuferin or salesperson or Einzelhandel or Verkauf or kaufmann or kauffrau or verkaufsberater or kundenberater or verkaufsberaterin or kundenberaterin'
        location = 'Buxtehude'
        resume_last_updated_days = 30
        target_candidates = 100
        max_radius = 50
    
    # Hardcoded parameters (not in Excel)
    language = "de"
    country = "DE"
    radius_increment = 5
    
    # Resume last updated filter (Unix timestamp in milliseconds)
    resume_updated_timestamp = calculate_unix_timestamp_ms(resume_last_updated_days)
    
    print(f"📋 Search Configuration:")
    print(f"   Keywords: {search_keywords}")
    print(f"   Location: {location}")
    print(f"   Target candidates: {target_candidates}")
    print(f"   Max radius: {max_radius}km")
    print(f"   Resume updated within: {resume_last_updated_days} days")
    
    # Get prompts from Excel parameters (configurable)
    user_prompt = params.get('user_prompt', 'EVALUATE each candidate against these EXACT criteria:...')
    system_prompt = params.get('system_prompt', 'You are an experienced recruiter specializing in fashion retail sales positions. Your task is to evaluate candidate CVs against specific criteria and return only the profile URLs of candidates who meet ALL requirements.')
    
    # Log search to Excel
    log_search_to_excel(search_id, search_keywords, location, f"Variable ({radius_increment}-{max_radius}km)", 
                       user_prompt=user_prompt, 
                       system_prompt=system_prompt)
    
    radius = radius_increment
    all_candidates = []
    
    while len(all_candidates) < target_candidates:
        print(f"\n=== SEARCHING RADIUS: {radius} km ===")
        print(f"Current total candidates: {len(all_candidates)}")
        print(f"Target: {target_candidates} candidates")
        
        # Use pagination to exhaust all candidates at this radius
        radius_candidates = fetch_all_candidates_with_pagination(
            radius, search_keywords, resume_updated_timestamp, 
            location, language, country, target_candidates, params
        )
        
        if radius_candidates:
            # Remove duplicates by checking profile URLs
            existing_urls = {candidate.get('profileLink', '') for candidate in all_candidates}
            new_candidates = [
                candidate for candidate in radius_candidates 
                if candidate.get('profileLink', '') not in existing_urls
            ]
            
            all_candidates.extend(new_candidates)
            print(f"Added {len(new_candidates)} new candidates from {radius}km radius")
            print(f"Total candidates now: {len(all_candidates)}")
            
            # If we reached our target, break
            if len(all_candidates) >= target_candidates:
                print(f"✅ Target of {target_candidates} candidates reached!")
                break
        else:
            print(f"No candidates found at {radius}km radius")
        
        # Increase radius and continue searching
        radius += radius_increment
        print(f"Expanding search radius to {radius} km...")
        
        # Safety limit to prevent infinite expansion
        if radius > max_radius:
            print(f"Maximum radius reached ({max_radius}km)")
            break
    
    candidates = all_candidates
    
    if candidates:
        # Create timestamp for filename
        timestamp = datetime.now().strftime("%d_%m_%Y_%H_%M_%S")
        filename = f"response_{timestamp}.json"
        
        # Save only the filtered data
        with open(filename, 'w') as f:
            json.dump(candidates, f, indent=2)
        
        print(f"Filtered data saved to: {filename}")
        print(f"Final result: {len(candidates)} candidates with {radius} km radius")
        
        # Automatically run MistralAI filtering
        print("\nStarting MistralAI filtering...")
        filter_candidates_with_mistral(search_keywords, location, radius, params)
    else:
        print("No candidates found")